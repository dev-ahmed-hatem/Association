from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.http import FileResponse
from rest_framework import viewsets, status

from financials.models import TransactionType
from .resources import fieldLabels
from .serializers import ProjectSerializer, ProjectTransactionReadSerializer, ProjectTransactionWriteSerializer
from rest_framework.response import Response
from rest_framework.decorators import action, api_view
from .models import Project, ProjectTransaction
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, RestrictedError, When, Case, F, DecimalField, ExpressionWrapper


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer

    def get_queryset(self):
        queryset = Project.objects.all()

        search = self.request.query_params.get('search', None)
        status_filters = self.request.query_params.get('status', None)
        from_date_str = self.request.query_params.get('from_date', None)
        to_date_str = self.request.query_params.get('to_date', None)
        sort_by = self.request.query_params.get('sort_by', None)
        order = self.request.query_params.get('order', None)

        if search is not None:
            queryset = queryset.filter(name__icontains=search)

        if status_filters:
            filters = status_filters.split(",")
            queryset = queryset.filter(status__in=filters)

        if from_date_str and to_date_str:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')

            queryset = queryset.filter(start_date__range=[from_date, to_date])

        if sort_by is not None:
            queryset = queryset.order_by(f"{order}{sort_by}")

        return queryset

    @action(detail=True, methods=['post'])
    def switch_status(self, request, pk=None):
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': _('مشروع غير موجود')}, status=status.HTTP_404_NOT_FOUND)
        new_status = request.data.get('status')

        project.status = new_status
        project.save()
        return Response({'status': project.get_status_display()}, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        try:
            return super(ProjectViewSet, self).destroy(request, *args, **kwargs)
        except RestrictedError:
            return Response(
                {"detail": _("لا يمكن حذف المشروع لارتباطه بسجلات مالية موجودة")},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def export_totals(self, request):
        queryset = self.get_queryset()
        fields = request.query_params.get("fields", None)

        if fields is None:
            return Response({"detail": "Please select fields to export"}, status=204)

        fields = fields.split(',')

        # totals annotation
        total_income = Sum(
            Case(
                When(
                    transactions__financial_record__transaction_type__type="إيراد",
                    then=F("transactions__financial_record__amount")
                ),
                output_field=DecimalField(),
                default=Decimal("0"),
            )
        )
        total_expense = Sum(
            Case(
                When(
                    transactions__financial_record__transaction_type__type="مصروف",
                    then=F("transactions__financial_record__amount")
                ),
                output_field=DecimalField(),
                default=Decimal("0")
            )
        )

        if "net_income" in fields:
            queryset = queryset.annotate(total_income=total_income, total_expense=total_expense).annotate(
                net_income=ExpressionWrapper(
                    F("total_income") - F("total_expense"),
                    output_field=DecimalField()
                ))
        else:
            annotations = {}
            if "total_income" in fields:
                annotations["total_income"] = total_income
            if "total_expense" in fields:
                annotations["total_expense"] = total_expense
            queryset = queryset.annotate(
                **annotations
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "المشاريع"

        ws.sheet_view.rightToLeft = True

        # Write fields
        for col, field in enumerate(fields, start=1):
            try:
                ws.cell(row=1, column=col, value=fieldLabels[field])
            except KeyError:
                pass

        # Write rows
        for row_num, item in enumerate(queryset, start=2):
            for col_num, field in enumerate(fields, start=1):
                if field == "created_by":
                    value = item.created_by.name if item.created_by else "-"
                elif field == "created_at":
                    value = item.created_at.astimezone(settings.CAIRO_TZ).strftime("%Y-%m-%d %I:%M%p")
                else:
                    value = getattr(item, field) or "-"
                ws.cell(row=row_num, column=col_num, value=value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="المشاريع.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @action(detail=False, methods=['get'])
    def export_monthly(self, request):
        queryset = self.get_queryset()
        fields = request.query_params.get("fields")
        start_date = request.query_params.get("start")
        end_date = request.query_params.get("end")

        if not fields:
            return Response({"detail": "Please select fields to export"}, status=204)

        fields = fields.split(',')

        if not start_date or not end_date:
            return Response({"detail": "Please provide 'start' and 'end' query params (YYYY-MM)"},
                            status=400)

        # Parse start and end months
        start = datetime.strptime(start_date, "%Y-%m")
        end = datetime.strptime(end_date, "%Y-%m")

        # Create list of months between start and end
        months = []
        current = start
        while current <= end:
            months.append(current)
            current += relativedelta(months=1)

        # Define income and expense annotations for a given month
        def month_annotations(selected_month):
            next_month = selected_month + relativedelta(months=1)
            return {
                'total_income': Sum(
                    Case(
                        When(
                            transactions__financial_record__transaction_type__type="إيراد",
                            transactions__financial_record__date__gte=selected_month,
                            transactions__financial_record__date__lt=next_month,
                            then=F("transactions__financial_record__amount")
                        ),
                        default=Decimal("0"),
                        output_field=DecimalField()
                    )
                ),
                'total_expense': Sum(
                    Case(
                        When(
                            transactions__financial_record__transaction_type__type="مصروف",
                            transactions__financial_record__date__gte=selected_month,
                            transactions__financial_record__date__lt=next_month,
                            then=F("transactions__financial_record__amount")
                        ),
                        default=Decimal("0"),
                        output_field=DecimalField()
                    )
                )
            }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التقارير الشهرية"
        ws.sheet_view.rightToLeft = True

        # Prepare header row
        headers = ["المشروع"]
        for month in months:
            month_name = month.strftime("%m-%Y")
            for f in fields:
                headers.append(f"{fieldLabels[f]} - {month_name}")

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Write project rows
        for row_num, item in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=str(item))  # project name or __str__

            col_offset = 2
            for month in months:
                ann = month_annotations(month)
                totals = queryset.filter(pk=item.pk).aggregate(**ann)
                income = totals["total_income"] or Decimal("0")
                expense = totals["total_expense"] or Decimal("0")
                net_income = income - expense

                for f in fields:
                    if f == "total_income":
                        value = income
                    elif f == "total_expense":
                        value = expense
                    elif f == "net_income":
                        value = net_income
                    else:
                        value = "-"
                    ws.cell(row=row_num, column=col_offset, value=value)
                    col_offset += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="التقارير_الشهرية.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class ProjectTransactionViewSet(viewsets.ModelViewSet):
    queryset = ProjectTransaction.objects.all()
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()

        project_id = self.request.query_params.get("project")
        if project_id:
            qs = qs.filter(project_id=project_id)

        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Split into incomes & expenses
        incomes = queryset.filter(financial_record__transaction_type__type=TransactionType.Type.INCOME)
        expenses = queryset.filter(financial_record__transaction_type__type=TransactionType.Type.EXPENSE)

        # Serialize
        income_serializer = self.get_serializer(incomes, many=True)
        expense_serializer = self.get_serializer(expenses, many=True)

        # Totals
        total_incomes = incomes.aggregate(total=Sum("financial_record__amount"))["total"] or 0
        total_expenses = expenses.aggregate(total=Sum("financial_record__amount"))["total"] or 0

        response_data = {
            "incomes": {
                "transactions": income_serializer.data,
                "total": total_incomes,
            },
            "expenses": {
                "transactions": expense_serializer.data,
                "total": total_expenses,
            },
            "net": total_incomes - total_expenses,
        }

        return Response(response_data)

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProjectTransactionWriteSerializer
        return ProjectTransactionReadSerializer

    @action(detail=True, methods=['patch'])
    def update_amount(self, request, pk=None):
        try:
            transaction = ProjectTransaction.objects.get(id=pk)
            amount = request.data.get('amount')
            transaction.financial_record.amount = amount
            transaction.financial_record.save()
            return Response({'amount': amount}, status=status.HTTP_200_OK)
        except Exception:
            return Response({'detail': _('عملية غير موجودة')}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def get_projects_stats(request):
    projects = Project.objects.all()
    total_projects = projects.count()
    in_progress = projects.filter(status=Project.Status.IN_PROGRESS).count()
    completed = projects.filter(status=Project.Status.COMPLETED).count()
    total_incomes = sum(project.total_incomes for project in projects)
    total_expenses = sum(project.total_expenses for project in projects)
    net = total_incomes - total_expenses

    return Response({"total_projects": total_projects,
                     "in_progress": in_progress,
                     "completed": completed,
                     "total_incomes": total_incomes,
                     "total_expenses": total_expenses,
                     "net": net
                     }, status=status.HTTP_200_OK)
