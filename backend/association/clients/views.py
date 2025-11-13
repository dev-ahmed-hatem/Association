from decimal import Decimal

import openpyxl
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.db.models.fields import IntegerField
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework import status

from association.utils import clean_excel_name
from financials.models import Installment, Subscription, FinancialRecord, TransactionType, Loan
from .models import Client, WorkEntity, RankChoices
from .serializers import WorkEntitySerializer, ClientListSerializer, ClientReadSerializer, ClientWriteSerializer, \
    ClientSelectSerializer
from django.utils.translation import gettext_lazy as _
from django.db.models import RestrictedError, Count, Sum, Value, CharField, Q, ExpressionWrapper, F, Case, When
from django.db.models.functions import TruncMonth, ExtractYear, ExtractMonth, Concat

from datetime import datetime, date

# for excel file creation
from io import BytesIO
from django.http import FileResponse
from .resourses import fieldLabels


class WorkEntityViewSet(ModelViewSet):
    queryset = WorkEntity.objects.all()
    serializer_class = WorkEntitySerializer

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except RestrictedError:
            return Response(
                {"detail": _("لا يمكن حذف جهة العمل لارتباطها بأعضاء موجودين")},
                status=status.HTTP_400_BAD_REQUEST
            )


class ClientViewSet(ModelViewSet):
    queryset = Client.objects.all()

    def get_queryset(self):
        queryset = Client.objects.all()

        search: str = self.request.query_params.get('search', None)
        search_type = self.request.query_params.get('search_type', "name__icontains")

        status_filters = self.request.query_params.get('status', [])
        rank_filters = self.request.query_params.get('rank', [])
        graduation_year_filters = self.request.query_params.get('graduation_year', [])
        entities_filters = self.request.query_params.get('entities', [])
        sort_by = self.request.query_params.get('sort_by', None)
        order = self.request.query_params.get('order', None)

        if search not in (None, ""):
            try:
                if search_type == "membership_number" and not search.isdigit():
                    raise ValueError("membership_number must be an integer")
                queryset = queryset.filter(**{search_type: search})
            except ValueError:
                pass

        if status_filters == "active":
            queryset = queryset.filter(is_active=True)
        elif status_filters == "retired":
            queryset = queryset.filter(is_active=False)

        if len(entities_filters) > 0:
            entities_filters = entities_filters.split(',')
            queryset = queryset.filter(work_entity__name__in=entities_filters)

        if len(rank_filters) > 0:
            rank_filters = rank_filters.split(',')
            queryset = queryset.filter(rank__in=rank_filters)

        if len(graduation_year_filters) > 0:
            graduation_year_filters = graduation_year_filters.split(',')
            queryset = queryset.filter(graduation_year__in=graduation_year_filters)

        if sort_by is not None:
            queryset = queryset.order_by(f"{order}{sort_by}")

        return queryset

    def get_serializer_class(self):
        serializer_type = self.request.query_params.get("serializer")
        if serializer_type == "select":
            return ClientSelectSerializer

        if self.action in ["create", "update", "partial_update"]:
            return ClientWriteSerializer
        return ClientListSerializer

    @action(detail=True, methods=['get'])
    def detailed(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk)
            data = ClientReadSerializer(client, context={"request": self.request}).data
            return Response(data)
        except Client.DoesNotExist:
            return Response({'detail': _('عضو غير موجود')}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def switch_active(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk)
            client.is_active = not client.is_active
            client.save()
            return Response({"is_active": client.is_active})
        except Exception:
            return Response({'detail': _('عضو غير موجود')}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'])
    def form_data(self, request, pk=None):
        try:
            client = Client.objects.get(id=pk)
            serializer = ClientWriteSerializer(client, context={"request": self.request}).data
            return Response(serializer)
        except Exception:
            return Response({'detail': _('عضو غير موجود')}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['delete'])
    def delete_financial_records(self, request, pk=None):
        try:
            client = Client.objects.get(id=pk)
            Installment.objects.filter(client=client).delete()
            Subscription.objects.filter(client=client).delete()
            Loan.objects.filter(client=client).delete()

            return Response(status=status.HTTP_204_NO_CONTENT)

        except Exception:
            return Response({'detail': _('عضو غير موجود')}, status=status.HTTP_404_NOT_FOUND)

    def destroy(self, request, *args, **kwargs):
        try:
            return super(ClientViewSet, self).destroy(request, *args, **kwargs)
        except RestrictedError:
            return Response(
                {"detail": _("لا يمكن حذف العميل لارتباطه بسجلات مالية موجودة")},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        queryset = self.get_queryset()
        fields = request.query_params.get("fields", None)

        if fields is None:
            return Response({"detail": "Please select fields to export"}, status=204)

        fields = fields.split(',')

        serializer = ClientReadSerializer(queryset, many=True, context={"request": request})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأعضاء"

        ws.sheet_view.rightToLeft = True

        data = serializer.data
        if not data:
            return Response({"detail": "No data to export"}, status=204)

        # Write fields
        for col, field in enumerate(fields, start=1):
            try:
                ws.cell(row=1, column=col, value=fieldLabels[field])
            except KeyError:
                pass

        # Write rows
        for row_num, item in enumerate(data, start=2):
            for col_num, field in enumerate(fields, start=1):
                if field == "is_active":
                    value = "في الخدمة" if item.get("is_active") else "متقاعد"
                else:
                    value = str(item.get(field, "") or "-")
                ws.cell(row=row_num, column=col_num, value=value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="الأعضاء.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @action(detail=True, methods=["get"])
    def export_subscriptions(self, request, pk=None):
        year = request.query_params.get("year", None)

        if not year:
            return Response({"detail": "Select Year"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
        except ValueError:
            return Response({"detail": "Invalid year format"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({"detail": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        subscriptions = Subscription.objects.filter(client=client, date__year=year).annotate(
            month=ExtractMonth("date")).values_list("month", "amount", "paid_at", "notes")

        months_dict = {
            month: {
                "amount": amount or Decimal("0.00"),
                "paid_at": paid_at,
                "notes": notes or "",
            }
            for month, amount, paid_at, notes in subscriptions
        }

        total_months = []
        for month in range(1, 13):
            record = months_dict.get(month, {
                "amount": Decimal("0.00"),
                "paid_at": "-",
                "notes": "-",
            })
            total_months.append({
                "month": f"{year}-{month}",
                **record,
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = clean_excel_name(f"سجل اشتراكات سنة {year} - {client.name}")
        ws.sheet_view.rightToLeft = True

        ws.cell(row=1, column=1, value="الشهر")
        ws.cell(row=1, column=2, value="المبلغ")
        ws.cell(row=1, column=3, value="تاريخ الدفع")
        ws.cell(row=1, column=4, value="ملاحظات")

        # Write rows
        for row_num, item in enumerate(total_months, start=2):
            ws.cell(row=row_num, column=1, value=item["month"])
            ws.cell(row=row_num, column=2, value=str(float(item["amount"])))
            ws.cell(row=row_num, column=3, value=item["paid_at"])
            ws.cell(row=row_num, column=4, value=item["notes"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename=f"سجل اشتراكات سنة {year} - {client.name}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @action(detail=True, methods=["get"])
    def export_installments(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({"detail": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        # 🔹 Query all installments for the given client and year
        installments = (
            Installment.objects.filter(client=client)
            .values_list("due_date", "amount", "status", "paid_at", "notes")
            .order_by("due_date")
        )

        # 🔹 Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = clean_excel_name(f"سجل الأقساط - {client.name}")
        ws.sheet_view.rightToLeft = True

        # Headers
        ws.cell(row=1, column=1, value="تاريخ الاستحقاق")
        ws.cell(row=1, column=2, value="القيمة")
        ws.cell(row=1, column=3, value="الحالة")
        ws.cell(row=1, column=4, value="تاريخ الدفع")
        ws.cell(row=1, column=5, value="ملاحظات")

        # Data rows
        for row_num, (due_date, amount, installment_status, paid_at, notes) in enumerate(installments, start=2):
            ws.cell(row=row_num, column=1, value=due_date.strftime("%Y-%m"))
            ws.cell(row=row_num, column=2, value=amount if installment_status == Installment.Status.PAID else "-")
            ws.cell(row=row_num, column=3, value=installment_status)
            ws.cell(row=row_num, column=4, value=paid_at or "-")
            ws.cell(row=row_num, column=5, value=notes or "-")

        # Save to in-memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as downloadable file
        return FileResponse(
            output,
            as_attachment=True,
            filename=f"سجل الأقساط - {client.name}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@api_view(["GET"])
def get_home_stats(request):
    # Rank Count
    ranks = Client.objects.values("rank").annotate(total=Count("id"))
    rank_dict = {r["rank"]: r["total"] for r in ranks}

    all_rank_counts = [{"rank": choice.value, "العدد": rank_dict.get(choice.value, 0)} for choice in RankChoices]

    # Activity Status
    activity = Client.objects.values("is_active").annotate(total=Count("id"))
    active_dict = {a["is_active"]: a["total"] for a in activity}

    active_status = [
        {"name": "بالخدمة", "value": active_dict.get(True, 0)},
        {"name": "متقاعد", "value": active_dict.get(False, 0)},
    ]

    entities_count = WorkEntity.objects.annotate(count=Count("client")).values("name", "id", "count")

    # Subscription Growth
    today = datetime.today().astimezone(settings.CAIRO_TZ).date()
    start_date = (today - relativedelta(months=6)).replace(day=1)

    month_totals = (Subscription.objects.filter(date__gte=start_date)
                    .annotate(month=TruncMonth("date"))
                    .annotate(
        month=Concat(ExtractYear("month"), Value("-"), ExtractMonth("month"), output_field=CharField(), ))
                    .values(
        "month").annotate(اشتراكات=Sum("amount")))

    data = {
        "rank_counts": all_rank_counts,
        "active_status": active_status,
        "entities_count": entities_count,
        "month_totals": month_totals,
    }

    return Response(data, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_home_financial_stats(request):
    now = datetime.now().astimezone(settings.CAIRO_TZ)
    current_month, current_year = now.month, now.year

    # ---------- Current month stats ----------
    incomes = (
            FinancialRecord.objects.filter(
                date__month=current_month,
                date__year=current_year,
                transaction_type__type=TransactionType.Type.INCOME,
            ).aggregate(total=Sum("amount"))["total"]
            or 0
    )
    expenses = (
            FinancialRecord.objects.filter(
                date__month=current_month,
                date__year=current_year,
                transaction_type__type=TransactionType.Type.EXPENSE,
            ).aggregate(total=Sum("amount"))["total"]
            or 0
    )

    # Subscriptions (current month)
    current_month_subs = Subscription.objects.filter(
        date__month=current_month, date__year=current_year, client__is_active=True
    )
    subscriptions_sum = current_month_subs.aggregate(total=Sum("amount"))["total"] or 0
    subscriptions_count = current_month_subs.count()

    cutoff = date(current_year, current_month, 1)
    unpaid_subscriptions = max(Client.objects.filter(is_active=True,
                                                     subscription_date__lt=cutoff).count() - subscriptions_count, 0)

    # Installments (current month)
    current_month_installments = Installment.objects.filter(
        due_date__month=current_month, due_date__year=current_year
    )
    installments_sum = (
            current_month_installments.filter(status=Installment.Status.PAID).aggregate(
                total=Sum("amount")
            )["total"]
            or 0
    )
    installments_count = current_month_installments.filter(
        status=Installment.Status.PAID
    ).count()
    unpaid_installments = current_month_installments.filter(
        status=Installment.Status.UNPAID
    ).count()

    # Loans (current month)
    loans_sum = Loan.objects.filter(
        issued_date__month=current_month, issued_date__year=current_year
    ).aggregate(total=Sum("amount"))["total"] or 0

    # ---------- Last 6 months ----------
    today = now.date()
    start_date = (today - relativedelta(months=6)).replace(day=1)
    end_date = (today.replace(day=1) + relativedelta(months=1))

    # Financial records (for charts, optional)
    monthly_totals = (
        FinancialRecord.objects.filter(date__gte=start_date)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(
            total_incomes=Sum(
                "amount", filter=Q(transaction_type__type=TransactionType.Type.INCOME)
            ),
            total_expenses=Sum(
                "amount", filter=Q(transaction_type__type=TransactionType.Type.EXPENSE)
            ),
        )
        .order_by("month")
    )

    subscription_stats = (Client.objects.filter(is_active=True)
    .annotate(
        start_year=ExtractYear("subscription_date"),
        start_month=ExtractMonth("subscription_date"),
    ).annotate(
        due_months=ExpressionWrapper(
            (current_year - F("start_year")) * 12 + (current_month - F("start_month")),
            output_field=IntegerField(),
        ),
        paid=Count("subscriptions"),
    ).annotate(
        unpaid=ExpressionWrapper(
            F("due_months") - F("paid"),
            output_field=IntegerField(),
        )
    )).aggregate(total_paid=Sum("paid"), total_unpaid=Sum(
        Case(
            When(unpaid__gt=0, then=F("unpaid")),
            default=Value(0),
            output_field=IntegerField(),
        )
    ))

    # Installments last 6 months (total unpaid like subscriptions)
    inst_till_now = Installment.objects.filter(due_date__lt=end_date).aggregate(
        paid=Count("id", filter=Q(status=Installment.Status.PAID)),
        unpaid=Count("id", filter=Q(status=Installment.Status.UNPAID)),
    )

    loans_data = Loan.objects.annotate(month=TruncMonth("issued_date")).values("month").annotate(value=Sum("amount"))

    return Response(
        {
            "month_totals": {
                "incomes": incomes,
                "expenses": expenses,
                "net": incomes - expenses,
                "subscriptions": subscriptions_sum,
                "installments": installments_sum,
                "loans": loans_sum
            },
            "last_6_monthly_totals": monthly_totals,
            "subscriptions_count": subscriptions_count,
            "unpaid_subscriptions": unpaid_subscriptions,
            "installments_count": installments_count,
            "unpaid_installments": unpaid_installments,
            "till_now_subs_inst": {
                "total_paid_subscriptions": subscription_stats["total_paid"],
                "total_paid_installments": inst_till_now["paid"] or 0,
                "total_unpaid_subscriptions": subscription_stats["total_unpaid"],
                "total_unpaid_installments": inst_till_now["unpaid"] or 0,
            },
            "loans_data": loans_data,
        },
        status=status.HTTP_200_OK,
    )
